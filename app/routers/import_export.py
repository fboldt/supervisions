import csv
from io import BytesIO, StringIO
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..auth import require_roles
from ..database import get_db
from ..models import Cohort, Role, StudentCurrent

router = APIRouter(prefix="/io", tags=["import-export"])


@router.get("/students/current.csv")
def export_current_csv(db: Session = Depends(get_db), _=Depends(require_roles(Role.admin, Role.staff))):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["registration_id", "full_name", "cohort_id", "advisor_1_id", "advisor_2_id", "scholarship"])
    for s in db.query(StudentCurrent).all():
        writer.writerow([s.registration_id, s.full_name, s.cohort_id, s.advisor_1_id, s.advisor_2_id, s.scholarship])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv")


@router.get("/cohorts.xlsx")
def export_cohorts_xlsx(db: Session = Depends(get_db), _=Depends(require_roles(Role.admin, Role.staff))):
    wb = Workbook()
    ws = wb.active
    ws.append(["code", "entrants", "defenses", "graduates", "publications_total"])
    for c in db.query(Cohort).all():
        ws.append([c.code, c.entrants, c.defenses, c.graduates, c.publications_total])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/students/current.xlsx")
def import_current_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_roles(Role.admin, Role.staff))):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only xlsx is supported")
    wb = load_workbook(BytesIO(file.file.read()))
    ws = wb.active
    imported = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            continue
        registration_id, full_name, cohort_id, advisor_1_id = row[:4]
        if not registration_id or not full_name:
            continue
        student = StudentCurrent(
            registration_id=str(registration_id),
            full_name=full_name,
            cohort_id=int(cohort_id),
            advisor_1_id=int(advisor_1_id),
        )
        db.add(student)
        imported += 1
    db.commit()
    return {"imported": imported}
