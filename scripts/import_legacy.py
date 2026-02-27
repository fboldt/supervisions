"""Legacy spreadsheet importer.
Usage:
  python scripts/import_legacy.py legacy.xlsx
"""
import sys
from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import Cohort


def main(path: str):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    db = SessionLocal()
    imported = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code, entrants, defenses, graduates, publications_total = row[:5]
        if not code:
            continue
        db.add(Cohort(code=str(code), entrants=int(entrants or 0), defenses=int(defenses or 0), graduates=int(graduates or 0), publications_total=int(publications_total or 0)))
        imported += 1
    db.commit()
    db.close()
    print(f"Imported {imported} cohorts")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("Provide path to xlsx file")
    main(sys.argv[1])
